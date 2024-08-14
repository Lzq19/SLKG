from transformers import AutoTokenizer, AutoModel, AutoConfig
import torch
import os
import json
from tqdm import tqdm
from enity_alignment import enity_alignment_e
import copy
import pandas as pd
from openpyxl import load_workbook, Workbook
from evl_re import acc_n,process_excel,count_missing_elements,RRC
from evl_entiy import sp
from F1 import calculate_f1


def model_act():
    ptuning_checkpoint = fr'./checkpoint'
    model_path = './thirdparty/glm/chatglm-6b'
    tokenizer = AutoTokenizer.from_pretrained(model_path, trust_remote_code=True)

    config = AutoConfig.from_pretrained(model_path, trust_remote_code=True, pre_seq_len=512)
    model = AutoModel.from_pretrained(model_path, config=config, trust_remote_code=True)
    prefix_state_dict = torch.load(os.path.join(ptuning_checkpoint, "pytorch_model.bin"))
    new_prefix_state_dict = {}
    for k, v in prefix_state_dict.items():
        if k.startswith("transformer.prefix_encoder."):
            new_prefix_state_dict[k[len("transformer.prefix_encoder."):]] = v
    model.transformer.prefix_encoder.load_state_dict(new_prefix_state_dict)
    model = model.quantize(4)
    model = model.half().cuda()
    model.transformer.prefix_encoder.float()
    model = model.eval()
    return model,tokenizer


if __name__ == "__main__":
    # predict
    file_path = './data/P-test.json'
    wb2 = Workbook() 
    ws2 = wb2.active

    data = []
    with open(file_path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            json_data = json.loads(line)
            content = json_data.get('content','')
            summary = json_data.get('summary','')
            p = summary.replace('，', ',').replace('；', ';')
            if p[-1] == ';':
                p = p[:-1]
            lines2 = [line2.strip() for line2 in p.split(';')]
            data2 = [line2.split(',') for line2 in lines2]
            for d2 in data2:
                ws2.append(d2)
            data.append(content)
            wb2.save("./data/label.xlsx")

    wb = Workbook() 
    ws = wb.active  

    wb3 = Workbook() 
    ws3 = wb3.active

    model,tokenizer = model_act()
    for query in tqdm(data):
        resp = model.chat(tokenizer, query, history=[])
        res = resp[0].replace('，', ',').replace('；', ';')
        if res[-1] == ';':
            res = res[:-1]
        lines = [line.strip() for line in res.split(';')]
        data = [line.split(',') for line in lines]

        data_e = copy.deepcopy(data)
        df = pd.read_csv('./data/bodyparts.csv', encoding='utf-8')
        bodyparts_dict = df.iloc[:, -1].tolist()
        for de in data_e:
            flag = de[1]
            if de[1] in bodyparts_dict:
                continue
            else:
                de[1]=enity_alignment_e(de[1])
            ws3.append(de)

        for d in data:
            ws.append(d)  

    wb.save("./data/SignKG-no-e.xlsx")
    wb3.save("./data/SignKG-e.xlsx")


    # 加载Excel工作簿  
    labelbook = load_workbook(rf'./data/label.xlsx')  
    Signbook = load_workbook(rf'./data/SignKG-no-e.xlsx')
    Signebook = load_workbook(rf'./data/SignKG-e.xlsx')
    # 选择第一个工作表（你也可以通过名字选择工作表）  
    worksheet_l = labelbook.active  
    worksheet_n = Signbook.active 
    worksheet_e = Signebook.active 
    # 初始化一个空列表，用于存放每一行的列表  
    label = []
    Sign_n = []
    Sign_e = []
    # 遍历工作表中的每一行（从第二行开始，因为第一行通常是标题行）  
    for row in worksheet_l.iter_rows(min_row=1, values_only=True):  
        # 将每一行的值转换为一个列表，并添加到rows_as_lists中  
        row_as_list = list(row)  
        label.append(row_as_list)  

    for row in worksheet_n.iter_rows(min_row=1, values_only=True):  
        # 将每一行的值转换为一个列表，并添加到rows_as_lists中  
        row_as_list = list(row)  
        Sign_n.append(row_as_list) 

    for row in worksheet_e.iter_rows(min_row=1, values_only=True):  
        # 将每一行的值转换为一个列表，并添加到rows_as_lists中  
        row_as_list = list(row)  
        Sign_e.append(row_as_list) 

    lbg = sp(label)
    Sng = sp(Sign_n)
    Seg = sp(Sign_e)

    right_n=right_e=fn=0
    for key in Sng:
        for t in Sng[key]:
            t = list(filter(None.__ne__, t))
            if len(t) != 4:
                fn+=1
            for sublist in lbg[key]:
                bodyp=t[1].replace(" ", "")
                lab = sublist[1].replace(" ", "")
                if bodyp == lab:
                    right_n+=1
                    break

    for key in Seg:
        for t1 in Seg[key]:
            for sublist in lbg[key]:
                bodyp=t1[1].replace(" ", "")
                lab = sublist[1].replace(" ", "")
                if bodyp == lab:
                    right_e+=1
                    break

    recall = 1-(fn/(len(label)-fn))
    acc_n = right_n/len(label)
    acc_e = right_e/len(label)
    print(f"Entity Recall: {recall:.4f}")
    print(f"Entity Accuracy (Normal): {acc_n:.4f}")
    print(f"Entity Accuracy (Enhanced): {acc_e:.4f}")


    file_paths = ['./data/label.xlsx', './data/SignKG-e.xlsx']  

    data_dict = {}  
    
    for file_path in file_paths:  
        data_dict[file_path] = process_excel(file_path)    

    recall_w = RRC(data_dict)
    rrecall = 1-(recall_w/(len(data_dict['./data/label.xlsx'])))

    
    label_sublists = data_dict['label.xlsx']
    sign_sublists = data_dict['SignKG-e.xlsx']
    worsen = acc_n(label_sublists,sign_sublists)
    accn = 1-(worsen/len(data_dict['label.xlsx']))

    worsee = count_missing_elements(data_dict)

    r_acc = 1 - worsee / len(data_dict['./data/label.xlsx'])
    print(f"Relation Recall: {rrecall:.4f}")
    print(f"Relation Accuracy: {r_acc:.4f}")


    # 计算F1分数  
    f1_score_n = calculate_f1(acc_n, recall)  
    f1_score_e = calculate_f1(acc_e, recall)  
    
    # 输出F1分数  
    print(f"Entity F1 Score (Normal): {f1_score_n:.4f}")
    print(f"Entity F1 Score (Enhanced): {f1_score_e:.4f}")
