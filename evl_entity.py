from openpyxl import load_workbook  


def sp(label):
    label_groups = {}
    for sublist in label:
        key = sublist[0]
        if key not in label_groups:
            label_groups[key] = []
        label_groups[key].append(sublist)
    return label_groups

if __name__=='__main__':
    labelbook = load_workbook(rf'./data/label.xlsx')  
    Signebook = load_workbook(rf'./data/SignKG-e.xlsx')

    worksheet_l = labelbook.active  
    worksheet_e = Signebook.active 

    label = []
    Sign_e = []

    for row in worksheet_l.iter_rows(min_row=1, values_only=True):  
        row_as_list = list(row)  
        label.append(row_as_list)  

    for row in worksheet_e.iter_rows(min_row=1, values_only=True):  
        row_as_list = list(row)  
        Sign_e.append(row_as_list) 

    lbg = sp(label)
    Seg = sp(Sign_e)

    right_e=fn=0
    for key in Seg:
        for t1 in Seg[key]:
            for sublist in lbg[key]:
                bodyp=t1[1].replace(" ", "")
                lab = sublist[1].replace(" ", "")
                if bodyp == lab:
                    right_e+=1
                    break

    recall = 1-(fn/(len(label)-fn))
    acc_e = right_e/len(label)
    print(f"Entity Recall: {recall:.4f}")
    print(f"Entity Accuracy : {acc_e:.4f}")