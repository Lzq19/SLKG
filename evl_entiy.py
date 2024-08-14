from openpyxl import load_workbook  


def sp(label):
    # 使用字典来根据第一个字符串元素分组  
    label_groups = {}
    for sublist in label:
        key = sublist[0]  # 第一个字符串元素作为键  
        if key not in label_groups:
            label_groups[key] = []
        label_groups[key].append(sublist)  # 将整个子列表添加到对应的组中  
    return label_groups

if __name__=='__mian__':
    # 加载Excel工作簿  
    labelbook = load_workbook(rf'./label.xlsx')  
    Signbook = load_workbook(rf'./SignKG-no-e.xlsx')
    Signebook = load_workbook(rf'./SignKG-e.xlsx')
    # 选择第一个工作表（你也可以通过名字选择工作表）  
    worksheet_l = labelbook.active  
    worksheet_n = Signbook.active 
    worksheet_e = Signebook.active 
    # 初始化一个空列表，用于存放每一行的列表  
    label = []
    Sign_n = []
    Sign_e = []
    # 遍历工作表中的每一行（从第二行开始，因为第一行通常是标题行）  
    for row in worksheet_l.iter_rows(min_row=1, values_only=True):  
        # 将每一行的值转换为一个列表，并添加到rows_as_lists中  
        row_as_list = list(row)  
        label.append(row_as_list)  

    for row in worksheet_n.iter_rows(min_row=1, values_only=True):  
        # 将每一行的值转换为一个列表，并添加到rows_as_lists中  
        row_as_list = list(row)  
        Sign_n.append(row_as_list) 

    for row in worksheet_e.iter_rows(min_row=1, values_only=True):  
        # 将每一行的值转换为一个列表，并添加到rows_as_lists中  
        row_as_list = list(row)  
        Sign_e.append(row_as_list) 

    lbg = sp(label)
    Sng = sp(Sign_n)
    Seg = sp(Sign_e)

    right_n=right_e=fn=0
    for key in Sng:
        for t in Sng[key]:
            t = list(filter(None.__ne__, t))
            if len(t) != 4:
                fn+=1
            for sublist in lbg[key]:
                bodyp=t[1].replace(" ", "")
                lab = sublist[1].replace(" ", "")
                if bodyp == lab:
                    right_n+=1
                    break

    for key in Seg:
        for t1 in Seg[key]:
            for sublist in lbg[key]:
                bodyp=t1[1].replace(" ", "")
                lab = sublist[1].replace(" ", "")
                if bodyp == lab:
                    right_e+=1
                    break

    recall = 1-(fn/(len(label)-fn))
    acc_n = right_n/len(label)
    acc_e = right_e/len(label)
    print(f"Entity Recall: {recall:.4f}")
    print(f"Entity Accuracy (Normal): {acc_n:.4f}")
    print(f"Entity Accuracy (Enhanced): {acc_e:.4f}")